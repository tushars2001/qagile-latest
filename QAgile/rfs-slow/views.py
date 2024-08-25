import os
import pdb
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from . import models
from ..admin.models import get_all_person
from openpyxl import load_workbook, drawing
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from datetime import datetime
from copy import copy
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required


@login_required(login_url='/identity/')
def rfs_home(request):

    if request.method == 'POST' and 'delete' in request.POST and 'rfs_request_id' in request.POST:
        rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        models.delete_rfs(rfs_request_id)

    rfsdata = models.get_rfsdata()

    context_data = {

        'rfsdata': rfsdata
    }
    return render(request, 'home.html', context_data)


@login_required(login_url='/identity/')
def rfs(request):
    rfs_request_id = ''

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])

    if request.method == "POST":
        if 'via_view' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        else:
            rfs_request_id = models.update_rfs_request(request.POST)

    rfs_request_data = models.get_rfs_request(rfs_request_id)
    rfs_request_scope_data = models.get_rfs_request_scope(rfs_request_id)
    persons_all = get_all_person()
    approved_rfs = models.get_approved_rfs(rfs_request_id)
    domains_info = models.get_domains_info()
    persons = list(filter(lambda persons: str(persons['role_id']) in ('1', '3', '10', '7'),
                          persons_all))
    poms = list(
        filter(lambda poms: str(poms['role_id']) in '2', persons_all))
    print(persons)
    if not rfs_request_data:
        rfs_request_data = [{'rfs_request_id': '', 'created': '',
                             'domain_id': '', 'last_updated': '', 'last_updated_by': '',
                             'current_project_status': '', 'qa_spoc': '', 'kickoff_date': '', 'repository': '',
                             'repository_access': '', 'alm_jira_link': '', 'project_id': '',
                             'project_name': '', 'category': '', 'requester_name': '',
                             'requester_email': '', 'requester_role': '', 'pm': '', 'PoM': '',
                             'dir_it': '', 'project_poc': '', 'project_type': '', 'project_size': '',
                             'tentative_prep_start': '', 'tentative_prep_end': '', 'tentative_plan_start': '',
                             'tentative_plan_end': '', 'tentative_exec_start': '', 'tentative_exec_end': '',
                             'tentative_close_start': '', 'tentative_close_end': '', 'project_description': '',
                             'qa_coe_lead': '', 'confidential': '', 'confidential_alt_name': '',
                             'confidential_doc_share_method': '', 'pii_bsi': '', 'pii_bsi_description': '',
                             'scope_remote': '', 'scope_ticket_num': '', 'rfs_status': ''}]

    if not rfs_request_scope_data:
        rfs_request_scope_data = [{'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '', 'startdate': '',
                                   'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '',
                                   'seq': ''}, {'rfs_request_id': '', 'service': 'UAT Support', 'include': '',
                                                'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '',
                                                'seq': ''}]

    print(rfs_request_data)
    print(rfs_request_scope_data)

    context_data = {
        'rfs_request_data': rfs_request_data[0],
        'rfs_request_scope_data_1': rfs_request_scope_data[0],
        'rfs_request_scope_data_2': rfs_request_scope_data[1],
        'rfs_request_scope_data_3': rfs_request_scope_data[2],
        'rfs_request_scope_data_4': rfs_request_scope_data[3],
        'rfs_request_scope_data_5': rfs_request_scope_data[4],
        'rfs_request_scope_data_6': rfs_request_scope_data[5],
        'rfs_request_scope_data_7': rfs_request_scope_data[6],
        'persons': persons,
        'poms': poms,
        'approved_rfs': approved_rfs,
        'domains_info': domains_info,
        'page': {'name': 'Request For Service'}
    }

    return render(request, 'request-for-service.html', context_data)


@login_required(login_url='/identity/')
def resource_loading(request):
    rfs_request_id = ''

    if request.method == "POST":
        if 'via_view' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        else:
            rfs_request_id = models.update_rfs_resourceLoading(request.POST)

    if request.method == "GET":
        rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])

    rfs_resourceLoading_data = models.get_rfs_resourceLoading(rfs_request_id)
    rfs_request_data = models.get_rfs_request(rfs_request_id)
    approved_rfs = models.get_approved_rfs(rfs_request_id)
    if not rfs_resourceLoading_data['datahead']:
        rfs_resourceLoading_data['datahead'] = [{'rfs_request_id': '', 'weekstart': '', 'numofweeks': ''}]

    persons_all = get_all_person()
    persons = [
        {k: v for k, v in d.items() if k in ('vnid', 'first_name', 'last_name', 'rate', 'domain_name', 'role', 'loc')}
        for d in persons_all]
    context_data = {
        'datadetails': rfs_resourceLoading_data['datadetails'],
        'datahead': rfs_resourceLoading_data['datahead'][0],
        'static_tbl': models.get_res_load_static(rfs_request_id),
        'persons': persons,
        'rfs_request_data': rfs_request_data[0],
        'approved_rfs': approved_rfs,
        'page': {'name': 'Resource Loading'}
    }
    return render(request, 'resource_loading.html', context_data)


@login_required(login_url='/identity/')
def estimates(request):
    rfs_request_id = ''
    print(request.POST)

    if request.method == "POST":
        if 'via_view' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        else:
            rfs_request_id = models.update_rfs_estimates(request.POST)

    rfs_request_data = models.get_rfs_request(rfs_request_id)

    rfs_estimates = models.get_rfs_estimates(rfs_request_id)
    rfs_estimates_raci_head = models.get_rfs_estimates_raci_head(rfs_request_id)
    rfs_estimates_raci = models.get_rfs_estimates_raci(rfs_request_id)
    rfs_total_rl = models.get_rfs_totals(rfs_request_id)
    approved_rfs = models.get_approved_rfs(rfs_request_id)
    rfs_estimates_raci_defaults = models.get_rfs_estimates_raci_defaults()

    print(rfs_estimates_raci_head)
    print(rfs_estimates_raci_defaults)

    if not rfs_estimates_raci_head:
        rfs_estimates_raci_head = [{'recordid': '', 'rfs_request_Id': '', 'SITRole': '', 'SITScope': '', 'UATRole': '',
                                'UATScope': '', 'total_qa_comments': ''}]

    if not rfs_total_rl:
        rfs_total_rl = [{'rfs_request_id': rfs_request_id, 'rfs_total_hrs': 0, 'total_qa_cost': 0}]

    context_data = {
        'rfs_request_data': rfs_request_data[0],
        'page': {'name': 'Estimates'},
        'rfs_estimates': rfs_estimates,
        'rfs_estimates_raci_head': rfs_estimates_raci_head[0],
        'rfs_estimates_raci': rfs_estimates_raci,
        'rfs_total_rl': rfs_total_rl[0],
        'approved_rfs': approved_rfs,
        'rfs_estimates_raci_defaults': rfs_estimates_raci_defaults
    }

    return render(request, 'estimates.html', context_data)


@login_required(login_url='/identity/')
def change_status(request):
    print(request.POST)
    return "data"


@login_required(login_url='/identity/')
def rfs_approved(request):
    rfs_request_id = ''
    version = 0

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])
        if 'version' in request.GET:
            version = request.GET['version']
    if request.method == "POST":
        if 'rfs_request_id' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        if 'version' in request.POST:
            version = request.POST['version']
    rfs_request_data = models.get_approved_rfs_request(rfs_request_id, version)
    rfs_request_scope_data = models.get_approved_rfs_request_scope(rfs_request_id, version)
    persons_all = get_all_person()
    approved_rfs = models.get_approved_rfs(rfs_request_id)

    persons = list(filter(lambda persons: str(persons['role_id']) in ('1', '3', '10'),
                          persons_all))
    poms = list(
        filter(lambda poms: str(poms['role_id']) in '2', persons_all))

    if not rfs_request_data:
        rfs_request_data = [{'rfs_request_id': '', 'created': '',
                             'domain_id': '', 'last_updated': '', 'last_updated_by': '',
                             'current_project_status': '', 'qa_spoc': '', 'kickoff_date': '', 'repository': '',
                             'repository_access': '', 'alm_jira_link': '', 'project_id': '',
                             'project_name': '', 'category': '', 'requester_name': '',
                             'requester_email': '', 'requester_role': '', 'pm': '', 'PoM': '',
                             'dir_it': '', 'project_poc': '', 'project_type': '', 'project_size': '',
                             'tentative_prep_start': '', 'tentative_prep_end': '', 'tentative_plan_start': '',
                             'tentative_plan_end': '', 'tentative_exec_start': '', 'tentative_exec_end': '',
                             'tentative_close_start': '', 'tentative_close_end': '', 'project_description': '',
                             'qa_coe_lead': '', 'confidential': '', 'confidential_alt_name': '',
                             'confidential_doc_share_method': '', 'pii_bsi': '', 'pii_bsi_description': '',
                             'scope_remote': '', 'scope_ticket_num': '', 'rfs_status': ''}]

    if not rfs_request_scope_data:
        rfs_request_scope_data = [{'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '', 'startdate': '',
                                   'enddate': '', 'comment': '', 'scope_id': '', 'seq': ''},
                                  {'rfs_request_id': '', 'service': '', 'include': '',
                                   'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '',
                                   'seq': ''}, {'rfs_request_id': '', 'service': 'UAT Support', 'include': '',
                                                'startdate': '', 'enddate': '', 'comment': '', 'scope_id': '',
                                                'seq': ''}]

    context_data = {
        'rfs_request_data': rfs_request_data[0],
        'rfs_request_scope_data_1': rfs_request_scope_data[0],
        'rfs_request_scope_data_2': rfs_request_scope_data[1],
        'rfs_request_scope_data_3': rfs_request_scope_data[2],
        'rfs_request_scope_data_4': rfs_request_scope_data[3],
        'rfs_request_scope_data_5': rfs_request_scope_data[4],
        'rfs_request_scope_data_6': rfs_request_scope_data[5],
        'rfs_request_scope_data_7': rfs_request_scope_data[6],
        'persons': persons,
        'poms': poms,
        'approved_rfs': approved_rfs,
        'version': version,
        'page': {'name': 'Request For Service'}
    }
    return render(request, 'request-for-service-approved.html', context_data)


@login_required(login_url='/identity/')
def resource_loading_approved(request):
    rfs_request_id = ''
    version = 0

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])
        if 'version' in request.GET:
            version = request.GET['version']

    if request.method == "POST":
        if 'rfs_request_id' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
            print("got rfs: ", rfs_request_id)
        if 'version' in request.POST:
            version = request.POST['version']
    rfs_resourceLoading_data = models.get_approved_rfs_resourceLoading(rfs_request_id, version)
    rfs_request_data = models.get_approved_rfs_request(rfs_request_id, version)
    approved_rfs = models.get_approved_rfs(rfs_request_id)
    if not rfs_resourceLoading_data['datahead']:
        rfs_resourceLoading_data['datahead'] = [{'rfs_request_id': '', 'weekstart': '', 'numofweeks': ''}]
    persons_all = get_all_person()
    persons = [
        {k: v for k, v in d.items() if k in ('vnid', 'first_name', 'last_name', 'rate', 'domain_name', 'role', 'loc')}
        for d in persons_all]
    print(persons)
    print(rfs_resourceLoading_data)
    context_data = {
        'datadetails': rfs_resourceLoading_data['datadetails'],
        'datahead': rfs_resourceLoading_data['datahead'][0],
        'static_tbl': models.get_approved_res_load_static(rfs_request_id, version),
        'persons': persons,
        'rfs_request_data': rfs_request_data[0],
        'approved_rfs': approved_rfs,
        'version': version,
        'page': {'name': 'Resource Loading'}
    }
    return render(request, 'resource_loading-approved.html', context_data)


@login_required(login_url='/identity/')
def estimates_approved(request):
    rfs_request_id = ''
    version = 0

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])
        if 'version' in request.GET:
            version = request.GET['version']

    if request.method == "POST":
        if 'rfs_request_id' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        if 'version' in request.POST:
            version = request.POST['version']

    rfs_request_data = models.get_approved_rfs_request(rfs_request_id, version)
    rfs_estimates = models.get_approved_rfs_estimates(rfs_request_id, version)
    rfs_estimates_raci_head = models.get_approved_rfs_estimates_raci_head(rfs_request_id, version)
    rfs_estimates_raci = models.get_approved_rfs_estimates_raci(rfs_request_id, version)
    rfs_total_rl = models.get_approved_rfs_totals(rfs_request_id, version)
    approved_rfs = models.get_approved_rfs(rfs_request_id)

    if not rfs_estimates_raci_head:
        rfs_estimates_raci_head = [{'recordid': '', 'rfs_request_Id': '', 'SITRole': '', 'SITScope': '', 'UATRole': '',
                                'UATScope': '', 'total_qa_comments': ''}]

    if not rfs_total_rl:
        rfs_total_rl = [{'rfs_request_id': rfs_request_id, 'rfs_total_hrs': 0, 'total_qa_cost': 0}]

    context_data = {
        'rfs_request_data': rfs_request_data[0],
        'page': {'name': 'Estimates'},
        'rfs_estimates': rfs_estimates,
        'rfs_estimates_raci_head': rfs_estimates_raci_head[0],
        'rfs_estimates_raci': rfs_estimates_raci,
        'rfs_total_rl': rfs_total_rl[0],
        'approved_rfs': approved_rfs,
        'version': version
    }

    return render(request, 'estimates-approved.html', context_data)


@login_required(login_url='/identity/')
def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])
        if 'version' in request.GET:
            version = request.GET['version']
            return export_users_xls_approved(request)

    # Get RFS related info
    rfs_request_data = models.get_rfs_request(rfs_request_id)
    rfs_request_scope_data = models.get_rfs_request_scope(rfs_request_id)
    rfs_resourse_loading = models.get_rfs_resourceLoading(rfs_request_id)
    rfs_estimates = models.get_rfs_estimates(rfs_request_id)
    rfs_estimates_raci_head = models.get_rfs_estimates_raci_head(rfs_request_id)
    rfs_estimates_raci = models.get_rfs_estimates_raci(rfs_request_id)
    rfs_total_rl = models.get_rfs_totals(rfs_request_id)

    if not rfs_estimates_raci_head:
        rfs_estimates_raci_head = [{'recordid': '', 'rfs_request_Id': '', 'SITRole': '', 'SITScope': '', 'UATRole': '',
                                'UATScope': '', 'total_qa_comments': ''}]
    project_name = rfs_request_data[0]['project_name'] if rfs_request_data[0]['project_name'] else ' '
    file = rfs_request_data[0]['project_id'] + " " + project_name + " - Request for Service V4.5.xlsx"
    response['Content-Disposition'] = 'attachment; filename="' + file + '"'

    #print(rfs_request_data)
    #print(rfs_request_scope_data)
    #print(rfs_resourse_loading)
    #print(rfs_estimates)
    #print(rfs_estimates_raci_head)
    #print("rfs_estimates_raci------------->")
    #print(rfs_estimates_raci)
    #print("rfs_estimates_raci------------->")
    #print(rfs_total_rl)
    #print("rfs_total_rl")
    # os.path.join(settings.STATIC_ROOT, 'rfs_template.xlsx')
    template = load_workbook(os.path.join(settings.STATIC_ROOT, 'rfs_template.xlsx'), data_only=True)

    processws = template.worksheets[0]
    rfs = template.worksheets[1]
    rfs_estimates_ws = template.worksheets[2]
    rfs_res_load = template.worksheets[3]

    # Updating Process worksheet start
    # img = drawing.image.Image(os.path.join(settings.STATIC_ROOT, 'images/rfs_process.png'))
    # img.anchor = 'B7'
    # processws.add_image(img)
    # Updating Process worksheet end

    # Updating RFS worksheet start
    rfs['B7'] = rfs_request_data[0]['last_updated']
    rfs['B8'] = rfs_request_data[0]['last_updated_by']
    rfs['B9'] = rfs_request_data[0]['current_project_status']
    rfs['B10'] = rfs_request_data[0]['qa_spoc']

    rfs['B14'] = rfs_request_data[0]['project_id']
    rfs['B15'] = rfs_request_data[0]['project_name']
    rfs['B16'] = rfs_request_data[0]['category']

    rfs['B18'] = rfs_request_data[0]['requester_name']
    rfs['B20'] = rfs_request_data[0]['requester_role']
    rfs['B22'] = rfs_request_data[0]['pm']
    rfs['B24'] = rfs_request_data[0]['first_name'] + " " + rfs_request_data[0]['last_name']
    rfs['B26'] = rfs_request_data[0]['dir_it']
    rfs['B28'] = rfs_request_data[0]['project_type']
    rfs['B30'] = rfs_request_data[0]['project_size']

    rfs['E7'] = rfs_request_data[0]['kickoff_date']
    rfs['E8'] = rfs_request_data[0]['repository']
    rfs['E9'] = rfs_request_data[0]['repository_access']
    rfs['E10'] = rfs_request_data[0]['alm_jira_link']

    rfs['E15'] = datetime.strptime(str(rfs_request_data[0]['tentative_prep_start'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_prep_start'] else ''
    rfs['F15'] = datetime.strptime(str(rfs_request_data[0]['tentative_prep_end'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_prep_end'] else ''
    rfs['E16'] = datetime.strptime(str(rfs_request_data[0]['tentative_plan_start'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_plan_start'] else ''
    rfs['F16'] = datetime.strptime(str(rfs_request_data[0]['tentative_plan_end'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_plan_end'] else ''
    rfs['E17'] = datetime.strptime(str(rfs_request_data[0]['tentative_exec_start'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_exec_start'] else ''
    rfs['F17'] = datetime.strptime(str(rfs_request_data[0]['tentative_exec_end'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_exec_end'] else ''
    rfs['E18'] = datetime.strptime(str(rfs_request_data[0]['tentative_close_start'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_close_start'] else ''
    rfs['F18'] = datetime.strptime(str(rfs_request_data[0]['tentative_close_end'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_data[0]['tentative_close_end'] else ''

    rfs['D24'] = rfs_request_data[0]['project_description']
    # rfs['E30'] = rfs_request_data[0]['qa_coe_lead']

    rfs['B33'] = rfs_request_data[0]['confidential']
    rfs['B35'] = rfs_request_data[0]['confidential_alt_name']
    rfs['B36'] = rfs_request_data[0]['confidential_doc_share_method']
    rfs['E33'] = rfs_request_data[0]['pii_bsi']
    rfs['E34'] = rfs_request_data[0]['pii_bsi_description']

    rfs['B38'] = rfs_request_data[0]['scope_remote']
    rfs['E38'] = rfs_request_data[0]['scope_ticket_num']

    for i in range(len(rfs_request_scope_data)):
        if i == 0:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C42'] = 'Yes'
            rfs['D42'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E42'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F42'] = rfs_request_scope_data[i]['comment']

        if i == 1:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C43'] = 'Yes'
            rfs['D43'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E43'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F43'] = rfs_request_scope_data[i]['comment']

        if i == 2:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C44'] = 'Yes'
            rfs['D44'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E44'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F44'] = rfs_request_scope_data[i]['comment']

        if i == 3:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C45'] = 'Yes'
            rfs['D45'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E45'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F45'] = rfs_request_scope_data[i]['comment']

        if i == 4:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C46'] = 'Yes'
            rfs['D46'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E46'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F46'] = rfs_request_scope_data[i]['comment']

        if i == 5:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C47'] = 'Yes'
            rfs['D47'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E47'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F47'] = rfs_request_scope_data[i]['comment']

        if i == 6:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C48'] = 'Yes'
            rfs['D48'] = datetime.strptime(str(rfs_request_scope_data[i]['startdate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['startdate'] else ''
            rfs['E48'] = datetime.strptime(str(rfs_request_scope_data[i]['enddate'])[:10], '%Y-%m-%d').strftime('%m/%d/%Y') if rfs_request_scope_data[i]['enddate'] else ''
            rfs['F48'] = rfs_request_scope_data[i]['comment']
    # Updating RFS worksheet End

    # Updating resource loading worksheet start
    person_id = ''
    row = 4
    col = 5
    numofweeks = rfs_resourse_loading['datahead'][0]['numofweeks']
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'),
                         right=Side(style='thick'),
                         top=Side(style='thick'),
                         bottom=Side(style='thick'))

    # these two LOC is to copy formatting
    cell = rfs_res_load['B4']
    cell_name = rfs_res_load['A5']

    # commenting out as now using fixed structure
    # rfs_res_load.cell(row, (4 + numofweeks + 1), 'Total Hours')
    # rfs_res_load.cell(row, (4 + numofweeks + 2), 'Rate')
    # rfs_res_load.cell(row, (4 + numofweeks + 3), 'QA Cost Amount')
    #
    # rfs_res_load.cell(row, (4 + numofweeks + 1)).fill = copy(cell.fill)
    # rfs_res_load.cell(row, (4 + numofweeks + 1)).font = copy(cell.font)
    # rfs_res_load.cell(row, (4 + numofweeks + 1)).border = thin_border
    # rfs_res_load.cell(row, (4 + numofweeks + 2)).fill = copy(cell.fill)
    # rfs_res_load.cell(row, (4 + numofweeks + 2)).font = copy(cell.font)
    # rfs_res_load.cell(row, (4 + numofweeks + 2)).border = thin_border
    # rfs_res_load.cell(row, (4 + numofweeks + 3)).fill = copy(cell.fill)
    # rfs_res_load.cell(row, (4 + numofweeks + 3)).font = copy(cell.font)
    # rfs_res_load.cell(row, (4 + numofweeks + 3)).border = thin_border

    # print(rfs_resourse_loading['datadetails'])
    startCol = 5
    wk = rfs_resourse_loading['datadetails'][0]['week']
    wkdt = datetime.strptime(wk, '%Y-%m-%d')
    wk = wkdt.strftime("%d-%b-%Y")
    for idx in range(1, 112):
        print("comparing :" + str(wkdt) + " and " + str(rfs_res_load.cell(4, idx).value))
        if wkdt == rfs_res_load.cell(4, idx).value:
            startCol = idx
            break

    for idx in range(5, 112):
        if idx < startCol or idx >= startCol + numofweeks:
            rfs_res_load.column_dimensions[get_column_letter(idx)].hidden = True

    for i in range(len(rfs_resourse_loading['datadetails'])):
        if person_id != rfs_resourse_loading['datadetails'][i]['person_id']:
            row = row + 1
            col = startCol
            person_id = rfs_resourse_loading['datadetails'][i]['person_id']
            if i > 1:
                rfs_res_load.insert_rows(row + 1, amount=1)
                print("adding row after: " + str(row))
            rfs_res_load.cell(row, 113, '=sum(E' + str(row) + ':DH' + str(row) + ')')
            rfs_res_load.cell(row, 115, '=DI' + str(row) + '*DJ' + str(row))
            rfs_res_load.cell(row, 113).border = thin_border
            rfs_res_load.cell(row, 115).border = thin_border
        rfs_res_load['A' + str(row)] = rfs_resourse_loading['datadetails'][i]['name']
        print(i, range(len(rfs_resourse_loading['datadetails'])))
        rfs_res_load['A' + str(row)].fill = copy(cell_name.fill)
        rfs_res_load['A' + str(row)].font = copy(cell_name.font)

        rfs_res_load['B' + str(row)] = rfs_resourse_loading['datadetails'][i]['role']
        rfs_res_load['C' + str(row)] = rfs_resourse_loading['datadetails'][i]['loc']
        rfs_res_load['D' + str(row)] = rfs_resourse_loading['datadetails'][i]['domain_name']

        rfs_res_load['B' + str(row)].border = thin_border
        rfs_res_load['C' + str(row)].border = thin_border
        rfs_res_load['D' + str(row)].border = thin_border

        # wk = rfs_resourse_loading['datadetails'][i]['week']
        # wkdt = datetime.strptime(wk, '%Y-%m-%d')
        # wk = wkdt.strftime("%d-%b-%Y")

        rfs_res_load.cell(row, col, rfs_resourse_loading['datadetails'][i]['hrs'])
        rfs_res_load.cell(row, col).border = thin_border
        # Fixed structure
        # rfs_res_load.cell(row, (4 + numofweeks + 2), float(rfs_resourse_loading['datadetails'][i]['rate']))
        # rfs_res_load.cell(row, (4 + numofweeks + 2)).number_format = 'General'
        rfs_res_load.cell(row, 114, float(rfs_resourse_loading['datadetails'][i]['rate']))
        rfs_res_load.cell(row, 114).number_format = 'General'
        rfs_res_load.cell(row, 114).border = thin_border
        # currval = rfs_res_load.cell(row, (4 + numofweeks + 1)).value
        # if not isinstance(currval, int):
        #     currval = 0
        # else:
        #     currval = int(rfs_res_load.cell(row, (4 + numofweeks + 1)).value)
        # rfs_res_load.cell(row, (4 + numofweeks + 1), (currval + int(rfs_resourse_loading['datadetails'][i]['hrs'])))

        # total_hrs = float(rfs_res_load.cell(row, (4 + numofweeks + 1)).value)
        # rate = float(rfs_res_load.cell(row, (4 + numofweeks + 2)).value)
        # cost = total_hrs * rate
        # rfs_res_load.cell(row, (4 + numofweeks + 3), cost)


        # if rfs_res_load.cell(3, col).value == 'Total Hours':
        #   rfs_res_load.insert_cols(col)
        # rfs_res_load.cell(4, col, wk).fill = copy(cell.fill)
        # rfs_res_load.cell(4, col).font = copy(cell.font)
        # rfs_res_load.cell(4, col).border = thin_border
        # rfs_res_load.cell(row=row, column=col).border = thin_border
        # rfs_res_load.cell(row, (4 + numofweeks + 1)).border = thin_border
        # rfs_res_load.cell(row, (4 + numofweeks + 2)).border = thin_border
        # rfs_res_load.cell(row, (4 + numofweeks + 3)).border = thin_border
        col = col + 1

    rfs_res_load.insert_rows(row + 1, amount=1)

    for i in range(111):
        total_cell = rfs_res_load.cell(row=(row + 1), column=(5+i))
        col_letter = get_column_letter((5+i))
        if col_letter != 'DJ':
            total_cell.value = '=sum('+col_letter+'5:'+col_letter+str(row)+')'
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
    rfs_res_load.cell(row=(row + 1), column=4, value='Total').font = Font(bold=True)
    rfs_res_load.insert_rows(row + 2, amount=1)
    rfs_res_load.cell(row=(row + 2), column=4, value='Monthly').font = Font(bold=True)
    rfs_res_load.cell(row=(row + 2), column=4).fill = copy(cell.fill)
    rfs_res_load.cell(row=(row + 1), column=4).fill = copy(cell.fill)
    rfs_res_load.cell(row=(row + 2), column=4).font = copy(cell.font)
    rfs_res_load.cell(row=(row + 1), column=4).font = copy(cell.font)
    rfs_res_load.cell(row=(row + 2), column=4).border = thin_border
    rfs_res_load.cell(row=(row + 1), column=4).border = thin_border
    mc_start = mc_end = None
    for i in range(5, 112):
        val = rfs_res_load.cell(row=3, column=i).value

        if val and len(val):
            if not mc_start:
                mc_start = i
                continue
            if mc_start and (mc_end is None or mc_end < mc_start):
                mc_end = i-1
                # do merge
                rfs_res_load.cell(row=row + 2, column=mc_start).value = '=sum(' + get_column_letter(mc_start) + str((row+1)) + ':' + get_column_letter(mc_end) + str((row+1)) + ')'
                rfs_res_load.merge_cells(get_column_letter(mc_start) + str(row+2) + ':' + get_column_letter(mc_end) + str(row+2))
                rfs_res_load.cell(row=row + 2, column=mc_start).alignment = Alignment(horizontal='center')
                mc_start = i
        rfs_res_load.cell(row=row + 2, column=mc_start).border = thick_border

    # Updating resource loading worksheet end

    # Updating resource loading worksheet start
    print("rfs_estimates-------------------------------------->")
    print(rfs_estimates)
    if rfs_estimates:
        rfs_estimates_ws['D8'] = rfs_estimates[0]['effort']
        rfs_estimates_ws['F8'] = rfs_estimates[0]['comment']
        rfs_estimates_ws['D9'] = rfs_estimates[1]['effort']
        rfs_estimates_ws['F9'] = rfs_estimates[1]['comment']
        rfs_estimates_ws['D10'] = '=sum(D8:D9)'

        rfs_estimates_ws['D13'] = rfs_estimates[2]['effort']
        rfs_estimates_ws['F13'] = rfs_estimates[2]['comment']
        rfs_estimates_ws['D14'] = rfs_estimates[3]['effort']
        rfs_estimates_ws['F14'] = rfs_estimates[3]['comment']
        rfs_estimates_ws['D15'] = rfs_estimates[4]['effort']
        rfs_estimates_ws['F15'] = rfs_estimates[4]['comment']
        rfs_estimates_ws['D16'] = rfs_estimates[5]['effort']
        rfs_estimates_ws['F16'] = rfs_estimates[5]['comment']
        rfs_estimates_ws['D17'] = rfs_estimates[6]['effort']
        rfs_estimates_ws['F17'] = rfs_estimates[6]['comment']
        rfs_estimates_ws['D18'] = rfs_estimates[7]['effort']
        rfs_estimates_ws['F18'] = rfs_estimates[7]['comment']
        rfs_estimates_ws['D19'] = rfs_estimates[8]['effort']
        rfs_estimates_ws['F19'] = rfs_estimates[8]['comment']
        rfs_estimates_ws['D20'] = '=sum(D13:D19)+D10'

        rfs_estimates_ws['D25'] = rfs_estimates[9]['effort']
        rfs_estimates_ws['F25'] = rfs_estimates[9]['comment']

        rfs_estimates_ws['D22'] = rfs_total_rl[0]['total_qa_cost'] + rfs_estimates[9]['effort']

        rfs_estimates_ws['D28'] = rfs_estimates_raci_head[0]['SITRole']
        rfs_estimates_ws['D50'] = rfs_estimates_raci_head[0]['UATRole']
        rfs_estimates_ws['D29'] = rfs_estimates_raci_head[0]['SITScope']
        rfs_estimates_ws['D51'] = rfs_estimates_raci_head[0]['UATScope']

        rfs_estimates_ws['D38'] = ''
        rfs_estimates_ws['E38'] = ''
        rfs_estimates_ws['D62'] = ''
        rfs_estimates_ws['E62'] = ''
        rfs_estimates_ws['D39'] = ''
        rfs_estimates_ws['E39'] = ''
        rfs_estimates_ws['D63'] = ''
        rfs_estimates_ws['E63'] = ''
        rfs_estimates_ws['D40'] = ''
        rfs_estimates_ws['E40'] = ''
        rfs_estimates_ws['D64'] = ''
        rfs_estimates_ws['E64'] = ''
        rfs_estimates_ws['D41'] = ''
        rfs_estimates_ws['E41'] = ''
        rfs_estimates_ws['D65'] = ''
        rfs_estimates_ws['E65'] = ''
        rfs_estimates_ws['D42'] = ''
        rfs_estimates_ws['E42'] = ''
        rfs_estimates_ws['D66'] = ''
        rfs_estimates_ws['E66'] = ''
        rfs_estimates_ws['D43'] = ''
        rfs_estimates_ws['E43'] = ''
        rfs_estimates_ws['D67'] = ''
        rfs_estimates_ws['E67'] = ''
        rfs_estimates_ws['D44'] = ''
        rfs_estimates_ws['E44'] = ''
        rfs_estimates_ws['D68'] = ''
        rfs_estimates_ws['E68'] = ''
        rfs_estimates_ws['D45'] = ''
        rfs_estimates_ws['E45'] = ''
        rfs_estimates_ws['D69'] = ''
        rfs_estimates_ws['E69'] = ''
        rfs_estimates_ws['D46'] = ''
        rfs_estimates_ws['E46'] = ''
        rfs_estimates_ws['D70'] = ''
        rfs_estimates_ws['E70'] = ''

        for i in range(len(rfs_estimates_raci)):
            if rfs_estimates_raci[i]['raci_code'] == 'CCQC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D38'] = str(rfs_estimates_ws['D38'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E38'] = str(rfs_estimates_ws['E38'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D62'] = str(rfs_estimates_ws['D62'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E62'] = str(rfs_estimates_ws['E62'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'TPLPR':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D39'] = str(rfs_estimates_ws['D39'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E39'] = str(rfs_estimates_ws['E39'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D63'] = str(rfs_estimates_ws['D63'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E63'] = str(rfs_estimates_ws['E63'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'TC2QC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D40'] = str(rfs_estimates_ws['D40'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E40'] = str(rfs_estimates_ws['E40'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D64'] = str(rfs_estimates_ws['D64'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E64'] = str(rfs_estimates_ws['E64'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'EXCTC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D41'] = str(rfs_estimates_ws['D41'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E41'] = str(rfs_estimates_ws['E41'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D65'] = str(rfs_estimates_ws['D65'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E65'] = str(rfs_estimates_ws['E65'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'MNGQC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D42'] = str(rfs_estimates_ws['D42'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E42'] = str(rfs_estimates_ws['E42'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D66'] = str(rfs_estimates_ws['D66'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E66'] = str(rfs_estimates_ws['E66'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'CDTER':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D43'] = str(rfs_estimates_ws['D43'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E43'] = str(rfs_estimates_ws['E43'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D67'] = str(rfs_estimates_ws['D67'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E67'] = str(rfs_estimates_ws['E67'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'RDTC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D44'] = str(rfs_estimates_ws['D44'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E44'] = str(rfs_estimates_ws['E44'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D68'] = str(rfs_estimates_ws['D68'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E68'] = str(rfs_estimates_ws['E68'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'WTR':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D45'] = str(rfs_estimates_ws['D45'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E45'] = str(rfs_estimates_ws['E45'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D69'] = str(rfs_estimates_ws['D69'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E69'] = str(rfs_estimates_ws['E69'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'MXE2E':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D46'] = str(rfs_estimates_ws['D46'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E46'] = str(rfs_estimates_ws['E46'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D70'] = str(rfs_estimates_ws['D70'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E70'] = str(rfs_estimates_ws['E70'].value) + ' ' + rfs_estimates_raci[i]['RACI']

    # Updating resource loading worksheet end

    template.save(response)
    return response


@login_required(login_url='/identity/')
def export_users_xls_approved(request):
    response = HttpResponse(content_type='application/ms-excel')

    rfs_request_id = ''
    version = 0

    if request.method == "GET":
        if 'rfs_request_id' in request.GET:
            rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])
        if 'version' in request.GET:
            version = request.GET['version']

    if request.method == "POST":
        if 'rfs_request_id' in request.POST:
            rfs_request_id = models.get_rfs_id_from_sha(request.POST['rfs_request_id'])
        if 'version' in request.POST:
            version = request.POST['version']

    rfs_request_id = rfs_request_id = models.get_rfs_id_from_sha(request.GET['rfs_request_id'])

    rfs_request_data = models.get_approved_rfs_request(rfs_request_id, version)
    rfs_request_scope_data = models.get_approved_rfs_request_scope(rfs_request_id, version)
    rfs_resourse_loading = models.get_approved_rfs_resourceLoading(rfs_request_id, version)
    rfs_estimates = models.get_approved_rfs_estimates(rfs_request_id, version)
    rfs_estimates_raci_head = models.get_approved_rfs_estimates_raci_head(rfs_request_id, version)
    rfs_estimates_raci = models.get_approved_rfs_estimates_raci(rfs_request_id, version)
    rfs_total_rl = models.get_approved_rfs_totals(rfs_request_id, version)

    if not rfs_estimates_raci_head:
        rfs_estimates_raci_head = [{'recordid': '', 'rfs_request_Id': '', 'SITRole': '', 'SITScope': '', 'UATRole': '',
                                'UATScope': '', 'total_qa_comments': ''}]

    project_name = rfs_request_data[0]['project_name'] if rfs_request_data[0]['project_name'] else ' '
    file = rfs_request_data[0]['project_id'] + " " + project_name + " - Request for Service V4.5.xlsx"
    response['Content-Disposition'] = 'attachment; filename="' + file + '"'

    #print(rfs_request_data)
    #print(rfs_request_scope_data)
    #print(rfs_resourse_loading)
    #print(rfs_estimates)
    #print(rfs_estimates_raci_head)
    print("rfs_estimates_raci------------->")
    print(rfs_estimates_raci)
    print("rfs_estimates_raci------------->")
    #print(rfs_total_rl)
    #print("rfs_total_rl")

    template = load_workbook(os.path.join(settings.STATIC_ROOT, 'rfs_template.xlsx'))

    processws = template.worksheets[0]
    rfs = template.worksheets[1]
    rfs_estimates_ws = template.worksheets[2]
    rfs_res_load = template.worksheets[3]

    # Updating Process worksheet start
    img = drawing.image.Image(
        os.path.join(settings.STATIC_ROOT, 'images/rfs_process.png'))
    img.anchor = 'B7'
    processws.add_image(img)
    # Updating Process worksheet end

    # Updating RFS worksheet start
    rfs['B7'] = rfs_request_data[0]['last_updated']
    rfs['B8'] = rfs_request_data[0]['last_updated_by']
    rfs['B9'] = rfs_request_data[0]['current_project_status']
    rfs['B10'] = rfs_request_data[0]['qa_spoc']

    rfs['B14'] = rfs_request_data[0]['project_id']
    rfs['B15'] = rfs_request_data[0]['project_name']
    rfs['B16'] = rfs_request_data[0]['category']

    rfs['B18'] = rfs_request_data[0]['requester_name']
    rfs['B20'] = rfs_request_data[0]['requester_role']
    rfs['B22'] = rfs_request_data[0]['pm']
    rfs['B24'] = rfs_request_data[0]['PoM']
    rfs['B26'] = rfs_request_data[0]['dir_it']
    rfs['B28'] = rfs_request_data[0]['project_type']
    rfs['B30'] = rfs_request_data[0]['project_size']

    rfs['E7'] = rfs_request_data[0]['kickoff_date']
    rfs['E8'] = rfs_request_data[0]['repository']
    rfs['E9'] = rfs_request_data[0]['repository_access']
    rfs['E10'] = rfs_request_data[0]['alm_jira_link']

    rfs['E15'] = rfs_request_data[0]['tentative_prep_start']
    rfs['F15'] = rfs_request_data[0]['tentative_prep_end']
    rfs['E16'] = rfs_request_data[0]['tentative_plan_start']
    rfs['F16'] = rfs_request_data[0]['tentative_plan_end']
    rfs['E17'] = rfs_request_data[0]['tentative_exec_start']
    rfs['F17'] = rfs_request_data[0]['tentative_exec_end']
    rfs['E18'] = rfs_request_data[0]['tentative_close_start']
    rfs['F18'] = rfs_request_data[0]['tentative_close_end']

    rfs['D24'] = rfs_request_data[0]['project_description']
    rfs['E30'] = rfs_request_data[0]['qa_coe_lead']

    rfs['B33'] = rfs_request_data[0]['confidential']
    rfs['B35'] = rfs_request_data[0]['confidential_alt_name']
    rfs['B36'] = rfs_request_data[0]['confidential_doc_share_method']
    rfs['E33'] = rfs_request_data[0]['pii_bsi']
    rfs['E34'] = rfs_request_data[0]['pii_bsi_description']

    rfs['B38'] = rfs_request_data[0]['scope_remote']
    rfs['E38'] = rfs_request_data[0]['scope_ticket_num']

    for i in range(len(rfs_request_scope_data)):
        if i == 0:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C42'] = 'Yes'
            rfs['D42'] = rfs_request_scope_data[i]['startdate']
            rfs['E42'] = rfs_request_scope_data[i]['enddate']
            rfs['F42'] = rfs_request_scope_data[i]['comment']

        if i == 1:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C43'] = 'Yes'
            rfs['D43'] = rfs_request_scope_data[i]['startdate']
            rfs['E43'] = rfs_request_scope_data[i]['enddate']
            rfs['F43'] = rfs_request_scope_data[i]['comment']

        if i == 2:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C44'] = 'Yes'
            rfs['D44'] = rfs_request_scope_data[i]['startdate']
            rfs['E44'] = rfs_request_scope_data[i]['enddate']
            rfs['F44'] = rfs_request_scope_data[i]['comment']

        if i == 3:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C45'] = 'Yes'
            rfs['D45'] = rfs_request_scope_data[i]['startdate']
            rfs['E45'] = rfs_request_scope_data[i]['enddate']
            rfs['F45'] = rfs_request_scope_data[i]['comment']

        if i == 4:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C46'] = 'Yes'
            rfs['D46'] = rfs_request_scope_data[i]['startdate']
            rfs['E46'] = rfs_request_scope_data[i]['enddate']
            rfs['F46'] = rfs_request_scope_data[i]['comment']

        if i == 5:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C47'] = 'Yes'
            rfs['D47'] = rfs_request_scope_data[i]['startdate']
            rfs['E47'] = rfs_request_scope_data[i]['enddate']
            rfs['F47'] = rfs_request_scope_data[i]['comment']

        if i == 6:
            if rfs_request_scope_data[i]['include'] == 1:
                rfs['C48'] = 'Yes'
            rfs['D48'] = rfs_request_scope_data[i]['startdate']
            rfs['E48'] = rfs_request_scope_data[i]['enddate']
            rfs['F48'] = rfs_request_scope_data[i]['comment']
    # Updating RFS worksheet End

    # Updating resource loading worksheet start
    person_id = ''
    row = 5
    col = 5
    numofweeks = rfs_resourse_loading['datahead'][0]['numofweeks']
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    cell = rfs_res_load['B4']

    rfs_res_load.cell(row, (4 + numofweeks + 1), 'Total Hours')
    rfs_res_load.cell(row, (4 + numofweeks + 2), 'Rate')
    rfs_res_load.cell(row, (4 + numofweeks + 3), 'QA Cost Amount')

    rfs_res_load.cell(row, (4 + numofweeks + 1)).fill = copy(cell.fill)
    rfs_res_load.cell(row, (4 + numofweeks + 1)).font = copy(cell.font)
    rfs_res_load.cell(row, (4 + numofweeks + 1)).border = thin_border
    rfs_res_load.cell(row, (4 + numofweeks + 2)).fill = copy(cell.fill)
    rfs_res_load.cell(row, (4 + numofweeks + 2)).font = copy(cell.font)
    rfs_res_load.cell(row, (4 + numofweeks + 2)).border = thin_border
    rfs_res_load.cell(row, (4 + numofweeks + 3)).fill = copy(cell.fill)
    rfs_res_load.cell(row, (4 + numofweeks + 3)).font = copy(cell.font)
    rfs_res_load.cell(row, (4 + numofweeks + 3)).border = thin_border

    print(rfs_resourse_loading['datadetails'])
    for i in range(len(rfs_resourse_loading['datadetails'])):
        if person_id != rfs_resourse_loading['datadetails'][i]['person_id']:
            row = row + 1
            col = 5
            person_id = rfs_resourse_loading['datadetails'][i]['person_id']
        rfs_res_load['A' + str(row)] = rfs_resourse_loading['datadetails'][i]['name']
        rfs_res_load['B' + str(row)] = rfs_resourse_loading['datadetails'][i]['role']
        rfs_res_load['C' + str(row)] = rfs_resourse_loading['datadetails'][i]['loc']
        rfs_res_load['D' + str(row)] = rfs_resourse_loading['datadetails'][i]['domain_name']
        rfs_res_load.cell(row, col, rfs_resourse_loading['datadetails'][i]['hrs'])
        rfs_res_load.cell(row, (4 + numofweeks + 2), float(rfs_resourse_loading['datadetails'][i]['rate']))
        rfs_res_load.cell(row, (4 + numofweeks + 2)).number_format = 'General'

        currval = rfs_res_load.cell(row, (4 + numofweeks + 1)).value
        if not isinstance(currval, int):
            currval = 0
        else:
            currval = int(rfs_res_load.cell(row, (4 + numofweeks + 1)).value)
        rfs_res_load.cell(row, (4 + numofweeks + 1), (currval + int(rfs_resourse_loading['datadetails'][i]['hrs'])))

        total_hrs = int(rfs_res_load.cell(row, (4 + numofweeks + 1)).value)
        rate = int(rfs_res_load.cell(row, (4 + numofweeks + 2)).value)
        cost = total_hrs * rate
        rfs_res_load.cell(row, (4 + numofweeks + 3), cost)

        wk = rfs_resourse_loading['datadetails'][i]['week']
        wkdt = datetime.strptime(wk, '%Y-%m-%d')
        wk = wkdt.strftime("%d-%b-%Y")
        # if rfs_res_load.cell(3, col).value == 'Total Hours':
        #   rfs_res_load.insert_cols(col)
        rfs_res_load.cell(3, col, wk).fill = copy(cell.fill)
        rfs_res_load.cell(3, col).font = copy(cell.font)
        rfs_res_load.cell(3, col).border = thin_border
        rfs_res_load.cell(row=row, column=col).border = thin_border
        rfs_res_load.cell(row, (4 + numofweeks + 1)).border = thin_border
        rfs_res_load.cell(row, (4 + numofweeks + 2)).border = thin_border
        rfs_res_load.cell(row, (4 + numofweeks + 3)).border = thin_border

        col = col + 1

    for i in range((numofweeks + 3)):
        total_cell = rfs_res_load.cell(row=(row + 1), column=(5+i))
        col_letter = get_column_letter((5+i))
        total_cell.value = '=sum('+col_letter+'1:'+col_letter+str(row)+')'
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
    rfs_res_load.cell(row=(row + 1), column=4, value='Total').font = Font(bold=True)

    # Updating resource loading worksheet end

    # Updating resource loading worksheet start
    print("rfs_estimates-------------------------------------->")
    print(rfs_estimates)
    if rfs_estimates:
        rfs_estimates_ws['D8'] = rfs_estimates[0]['effort']
        rfs_estimates_ws['F8'] = rfs_estimates[0]['comment']
        rfs_estimates_ws['D9'] = rfs_estimates[1]['effort']
        rfs_estimates_ws['F9'] = rfs_estimates[1]['comment']
        rfs_estimates_ws['D10'] = '=sum(D8:D9)'

        rfs_estimates_ws['D13'] = rfs_estimates[2]['effort']
        rfs_estimates_ws['F13'] = rfs_estimates[2]['comment']
        rfs_estimates_ws['D14'] = rfs_estimates[3]['effort']
        rfs_estimates_ws['F14'] = rfs_estimates[3]['comment']
        rfs_estimates_ws['D15'] = rfs_estimates[4]['effort']
        rfs_estimates_ws['F15'] = rfs_estimates[4]['comment']
        rfs_estimates_ws['D16'] = rfs_estimates[5]['effort']
        rfs_estimates_ws['F16'] = rfs_estimates[5]['comment']
        rfs_estimates_ws['D17'] = rfs_estimates[6]['effort']
        rfs_estimates_ws['F17'] = rfs_estimates[6]['comment']
        rfs_estimates_ws['D18'] = rfs_estimates[7]['effort']
        rfs_estimates_ws['F18'] = rfs_estimates[7]['comment']
        rfs_estimates_ws['D19'] = rfs_estimates[8]['effort']
        rfs_estimates_ws['F19'] = rfs_estimates[8]['comment']
        rfs_estimates_ws['D20'] = '=sum(D13:D19)+D10'

        rfs_estimates_ws['D25'] = rfs_estimates[9]['effort']
        rfs_estimates_ws['F25'] = rfs_estimates[9]['comment']

        rfs_estimates_ws['D22'] = rfs_total_rl[0]['total_qa_cost'] + rfs_estimates[9]['effort']

        rfs_estimates_ws['D28'] = rfs_estimates_raci_head[0]['SITRole']
        rfs_estimates_ws['D50'] = rfs_estimates_raci_head[0]['UATRole']
        rfs_estimates_ws['D29'] = rfs_estimates_raci_head[0]['SITScope']
        rfs_estimates_ws['D51'] = rfs_estimates_raci_head[0]['UATScope']

        rfs_estimates_ws['D38'] = ''
        rfs_estimates_ws['E38'] = ''
        rfs_estimates_ws['D62'] = ''
        rfs_estimates_ws['E62'] = ''
        rfs_estimates_ws['D39'] = ''
        rfs_estimates_ws['E39'] = ''
        rfs_estimates_ws['D63'] = ''
        rfs_estimates_ws['E63'] = ''
        rfs_estimates_ws['D40'] = ''
        rfs_estimates_ws['E40'] = ''
        rfs_estimates_ws['D64'] = ''
        rfs_estimates_ws['E64'] = ''
        rfs_estimates_ws['D41'] = ''
        rfs_estimates_ws['E41'] = ''
        rfs_estimates_ws['D65'] = ''
        rfs_estimates_ws['E65'] = ''
        rfs_estimates_ws['D42'] = ''
        rfs_estimates_ws['E42'] = ''
        rfs_estimates_ws['D66'] = ''
        rfs_estimates_ws['E66'] = ''
        rfs_estimates_ws['D43'] = ''
        rfs_estimates_ws['E43'] = ''
        rfs_estimates_ws['D67'] = ''
        rfs_estimates_ws['E67'] = ''
        rfs_estimates_ws['D44'] = ''
        rfs_estimates_ws['E44'] = ''
        rfs_estimates_ws['D68'] = ''
        rfs_estimates_ws['E68'] = ''
        rfs_estimates_ws['D45'] = ''
        rfs_estimates_ws['E45'] = ''
        rfs_estimates_ws['D69'] = ''
        rfs_estimates_ws['E69'] = ''
        rfs_estimates_ws['D46'] = ''
        rfs_estimates_ws['E46'] = ''
        rfs_estimates_ws['D70'] = ''
        rfs_estimates_ws['E70'] = ''

        for i in range(len(rfs_estimates_raci)):
            if rfs_estimates_raci[i]['raci_code'] == 'CCQC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D38'] = str(rfs_estimates_ws['D38'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E38'] = str(rfs_estimates_ws['E38'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D62'] = str(rfs_estimates_ws['D62'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E62'] = str(rfs_estimates_ws['E62'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'TPLPR':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D39'] = str(rfs_estimates_ws['D39'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E39'] = str(rfs_estimates_ws['E39'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D63'] = str(rfs_estimates_ws['D63'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E63'] = str(rfs_estimates_ws['E63'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'TC2QC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D40'] = str(rfs_estimates_ws['D40'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E40'] = str(rfs_estimates_ws['E40'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D64'] = str(rfs_estimates_ws['D64'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E64'] = str(rfs_estimates_ws['E64'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'EXCTC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D41'] = str(rfs_estimates_ws['D41'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E41'] = str(rfs_estimates_ws['E41'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D65'] = str(rfs_estimates_ws['D65'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E65'] = str(rfs_estimates_ws['E65'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'MNGQC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D42'] = str(rfs_estimates_ws['D42'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E42'] = str(rfs_estimates_ws['E42'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D66'] = str(rfs_estimates_ws['D66'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E66'] = str(rfs_estimates_ws['E66'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'CDTER':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D43'] = str(rfs_estimates_ws['D43'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E43'] = str(rfs_estimates_ws['E43'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D67'] = str(rfs_estimates_ws['D67'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E67'] = str(rfs_estimates_ws['E67'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'RDTC':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D44'] = str(rfs_estimates_ws['D44'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E44'] = str(rfs_estimates_ws['E44'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D68'] = str(rfs_estimates_ws['D68'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E68'] = str(rfs_estimates_ws['E68'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'WTR':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D45'] = str(rfs_estimates_ws['D45'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E45'] = str(rfs_estimates_ws['E45'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D69'] = str(rfs_estimates_ws['D69'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E69'] = str(rfs_estimates_ws['E69'].value) + ' ' + rfs_estimates_raci[i]['RACI']

            if rfs_estimates_raci[i]['raci_code'] == 'MXE2E':
                if rfs_estimates_raci[i]['raci_scope'] == 'SIT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D46'] = str(rfs_estimates_ws['D46'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E46'] = str(rfs_estimates_ws['E46'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                if rfs_estimates_raci[i]['raci_scope'] == 'UAT':
                    if rfs_estimates_raci[i]['p_or_q'] == 'P':
                        rfs_estimates_ws['D70'] = str(rfs_estimates_ws['D70'].value) + ' ' + rfs_estimates_raci[i]['RACI']
                    if rfs_estimates_raci[i]['p_or_q'] == 'Q':
                        rfs_estimates_ws['E70'] = str(rfs_estimates_ws['E70'].value) + ' ' + rfs_estimates_raci[i]['RACI']

    # Updating resource loading worksheet end

    template.save(response)
    return response


def person_week(request):
    data = {'success': 'False', 'error': '', 'data': {}}

    if request.method == 'GET' and 'vnid' in request.GET and len('vnid') and 'week' in request.GET and len(request.GET['week']):
        data['data'] = models.get_person_week(request.GET['vnid'], request.GET['week'])
        total_hrs = 0
        data['success'] = 'True'
        for idx in range(len(data['data'])):
            total_hrs = total_hrs + float(data['data'][idx]['hrs'])
        data['summary'] = {'vnid': request.GET['vnid'], 'week': request.GET['week'], 'total_hrs': total_hrs}
    else:
        data['error'] = "Need valid vnid and week in yyyy-mm-dd."

    return JsonResponse(data)


def rfs_hrs_by_role(request):
    data = {'success': 'False', 'error': '', 'data': {}}

    if request.method == 'GET' and 'role_id' in request.GET and len('role_id') and 'link' in request.GET and len(
            request.GET['link']):
        rfs_request_id = models.get_rfs_id_from_sha(request.GET['link'])
        data['data'] = models.get_rfs_hrs_by_role(rfs_request_id, request.GET['role_id'])
        data['success'] = 'true'
        data['rfs_request_id'] = rfs_request_id
        data['role_id'] = request.GET['role_id']
        data['link'] = request.GET['link']
    else:
        data['error'] = "Need valid role_id and rfs link."

    return JsonResponse(data)


def get_period_info(date):
    info = {'date': date, 'period': None, 'period_start': None, 'period_end': None}

    return info


@csrf_exempt
def rfs_auditor(request):
    ret = {'success': False, 'error': None, 'results': None}
    rfs_request_ids = None
    if request.method == 'POST':
        if 'rfs_request_ids' in request.POST:
            rfs_request_ids = request.POST['rfs_request_ids']

        ret['results'] = models.rfs_auditor(rfs_request_ids)
        ret['success'] = True
    else:
        ret['error'] = 'Invalid request'

    return JsonResponse(ret)
